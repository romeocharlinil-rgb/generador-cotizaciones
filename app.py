from datetime import date
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill


st.set_page_config(
    page_title="Generador de Cotizaciones",
    page_icon=":memo:",
    layout="centered",
)


def formato_moneda(valor: float) -> str:
    return f"${valor:,.2f}"


def crear_excel(cotizacion: dict) -> bytes:
    archivo = BytesIO()

    resumen = pd.DataFrame(
        [
            ["Nombre del negocio", cotizacion["negocio"]],
            ["Nombre del cliente", cotizacion["cliente"]],
            ["Fecha", cotizacion["fecha"].strftime("%Y-%m-%d")],
            ["Observaciones", cotizacion["observaciones"]],
        ],
        columns=["Campo", "Valor"],
    )

    detalle = pd.DataFrame(
        [
            {
                "Producto o servicio": cotizacion["producto"],
                "Cantidad": cotizacion["cantidad"],
                "Precio unitario": cotizacion["precio_unitario"],
                "Subtotal": cotizacion["subtotal"],
                "Descuento": cotizacion["descuento_monto"],
                "Impuesto": cotizacion["impuesto_monto"],
                "Total final": cotizacion["total"],
            }
        ]
    )

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="Cotizacion", startrow=2)
        detalle.to_excel(writer, index=False, sheet_name="Cotizacion", startrow=9)

        hoja = writer.sheets["Cotizacion"]
        hoja["A1"] = "Generador de Cotizaciones"
        hoja["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        hoja["A1"].fill = PatternFill("solid", fgColor="2563EB")
        hoja["A1"].alignment = Alignment(horizontal="center")
        hoja.merge_cells("A1:G1")

        for fila in hoja.iter_rows(min_row=3, max_row=4, min_col=1, max_col=2):
            for celda in fila:
                celda.alignment = Alignment(wrap_text=True, vertical="top")

        for celda in hoja[10]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="172033")

        for columna in hoja.columns:
            ancho = max(len(str(celda.value)) if celda.value is not None else 0 for celda in columna)
            hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 4, 32)

    return archivo.getvalue()


st.markdown(
    """
    <style>
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .titulo {
        font-size: 2rem;
        font-weight: 800;
        color: #172033;
        margin-bottom: 0.25rem;
    }
    .subtitulo {
        color: #5b667a;
        margin-bottom: 1.5rem;
    }
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #e6eaf0;
        border-radius: 8px;
        padding: 0.9rem;
    }
    .vista-previa {
        border: 1px solid #e6eaf0;
        border-radius: 8px;
        padding: 1rem;
        background: #ffffff;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="titulo">Generador de Cotizaciones</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="subtitulo">Crea una cotizacion simple, calcula el total y descargala en Excel.</div>',
    unsafe_allow_html=True,
)

with st.container():
    st.subheader("Formulario")

    negocio = st.text_input("Nombre del negocio", value="Mi Negocio")
    cliente = st.text_input("Nombre del cliente", value="Cliente Ejemplo")
    fecha = st.date_input("Fecha", value=date.today())
    producto = st.text_input("Producto o servicio", value="Servicio profesional")

    col1, col2 = st.columns(2)
    with col1:
        cantidad = st.number_input("Cantidad", min_value=1, value=1, step=1)
    with col2:
        precio_unitario = st.number_input("Precio unitario", min_value=0.0, value=100.0, step=1.0)

    col3, col4 = st.columns(2)
    with col3:
        descuento_pct = st.number_input("Descuento opcional (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0)
    with col4:
        impuesto_pct = st.number_input("Impuesto opcional (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0)

    observaciones = st.text_area("Observaciones", value="Gracias por su preferencia.")

subtotal = cantidad * precio_unitario
descuento_monto = subtotal * (descuento_pct / 100)
base_impuesto = subtotal - descuento_monto
impuesto_monto = base_impuesto * (impuesto_pct / 100)
total = base_impuesto + impuesto_monto

cotizacion = {
    "negocio": negocio,
    "cliente": cliente,
    "fecha": fecha,
    "producto": producto,
    "cantidad": cantidad,
    "precio_unitario": precio_unitario,
    "descuento_pct": descuento_pct,
    "descuento_monto": descuento_monto,
    "impuesto_pct": impuesto_pct,
    "impuesto_monto": impuesto_monto,
    "subtotal": subtotal,
    "total": total,
    "observaciones": observaciones,
}

st.subheader("Vista previa")

st.markdown('<div class="vista-previa">', unsafe_allow_html=True)
st.write(f"**Negocio:** {negocio}")
st.write(f"**Cliente:** {cliente}")
st.write(f"**Fecha:** {fecha.strftime('%Y-%m-%d')}")
st.write(f"**Producto o servicio:** {producto}")
st.write(f"**Cantidad:** {cantidad}")
st.write(f"**Precio unitario:** {formato_moneda(precio_unitario)}")
st.write(f"**Observaciones:** {observaciones}")
st.markdown("</div>", unsafe_allow_html=True)

col5, col6, col7, col8 = st.columns(4)
col5.metric("Subtotal", formato_moneda(subtotal))
col6.metric("Descuento", formato_moneda(descuento_monto))
col7.metric("Impuesto", formato_moneda(impuesto_monto))
col8.metric("Total final", formato_moneda(total))

st.download_button(
    label="Descargar cotizacion en Excel",
    data=crear_excel(cotizacion),
    file_name="cotizacion.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
